"""Generate Excel spreadsheets from CNPJ query results."""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.core.cnpj_validator import format_cnpj
from src.database.models import CNPJQuery

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

COLUMNS = [
    ("CNPJ", 20),
    ("Nome Empresarial", 40),
    ("Situação Simples Nacional", 35),
    ("Situação SIMEI", 30),
    ("Periodos Anteriores (SN)", 35),
    ("Periodos Anteriores (SIMEI)", 35),
    ("Eventos Futuros (SN)", 35),
    ("Eventos Futuros (SIMEI)", 35),
    ("MEI Transportador Autonomo de Cargas", 35),
    ("Status", 14),
    ("Erro", 40),
    ("Data da Consulta", 22),
]


def generate_spreadsheet(
    queries: list[CNPJQuery],
    export_dir: str,
    request_id: str,
) -> Path:
    """Create an .xlsx file with the query results and return its path."""
    export_path = Path(export_dir)
    export_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"consulta_simples_{request_id[:8]}_{timestamp}.xlsx"
    filepath = export_path / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta Simples Nacional"

    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, q in enumerate(queries, start=2):
        consulted = (
            q.consulted_at.strftime("%d/%m/%Y %H:%M:%S") if q.consulted_at else ""
        )
        row_data = [
            format_cnpj(q.cnpj),
            q.nome_empresarial or "",
            q.situacao_simples or "",
            q.situacao_simei or "",
            q.periodos_anteriores_sn or "",
            q.periodos_anteriores_simei or "",
            q.eventos_futuros_sn or "",
            q.eventos_futuros_simei or "",
            q.mei_transportador_autonomo_cargas or "",
            q.status,
            q.error_message or "",
            consulted,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(str(filepath))
    logger.info("Spreadsheet saved: %s", filepath)
    return filepath
